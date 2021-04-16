"""
хочу найти яцейку с данной информацией в ней
и получить ее координаты - успешно
"""

from openpyxl import load_workbook
file = "/Volumes/big4photo/Documents/Tass_total_report.xlsx"
wb = load_workbook(filename=file)
ws = wb.active
for row in ws.iter_rows():
    for cell in row:
        if cell.value == 3970707:
            print(f'row = {cell.row}, column = {cell.column}')
