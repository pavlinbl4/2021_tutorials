from openpyxl import Workbook
import datetime
import calendar


month_number = 1  # int(input("введите номер месяца"))
name_of_month = calendar.month_name[month_number]


wb = Workbook()
ws = wb.active
ws.title = name_of_month
ws.sheet_properties.tabColor = "1072BA"
ws2 = wb.create_sheet("Mysheet", 0)
ws.cell(row =1,column=2).value = "Дата"
ws.cell(row =3,column=1).value = datetime.datetime.now()
wb.save("/Volumes/big4photo/Documents/Kommersant/test.xlsx")
wb.close()