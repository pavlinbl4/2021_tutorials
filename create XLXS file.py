import openpyxl
import calendar
import datetime
from openpyxl import load_workbook

# month_number = int(input("введите номер месяца"))
# name_of_month = calendar.month_name[month_number]


book = openpyxl.Workbook() # создание новой таблицы с перезаписью старой


# ws = book.create_sheet(name_of_month,0)


# book.title = "Kommersant"
# ws.cell(row =1,column=2).value = "Дата"
# ws.cell(row =3,column=1).value = datetime.datetime.now()
#
# ws.cell(row =2,column=1).value = "photo id"
# ws.cell(row =3,column=2).value = "7777"

book.save("/Volumes/big4photo/Documents/Kommersant/publications.xlsx")
book.close()