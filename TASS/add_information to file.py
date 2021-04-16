"""
проверка возможности дописывать в файл
"""

file = "/Volumes/big4photo/Documents/Tass_total_report.xlsx"
from openpyxl import load_workbook
wb = load_workbook(filename=file, read_only=False)
ws = wb.active

ws.cell(row =5,column=1).value = "add new info"

wb.save(file)