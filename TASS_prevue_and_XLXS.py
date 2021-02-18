# проходится по файлу отчета и скачивает превью изображений
from selenium import webdriver
import time
import requests
import openpyxl

options = webdriver.ChromeOptions()
options.add_argument("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4200.0 Iron Safari/537.36")
browser = webdriver.Chrome(options=options)

wb = openpyxl.load_workbook('/Users/evgeniy/Downloads/Павленко-2.xlsx')
sheet = wb.active
photos = {}
x = 7

try:
    browser.get('https://www.tassphoto.com/ru')
    time.sleep(1)

    photo_id = (sheet.cell(row=x, column=4)).value
    while photo_id != None:
        search_input = browser.find_element_by_id("userrequest")
        search_input.clear()
        search_input.send_keys(photo_id)
        browser.find_element_by_id("search-submit").click()
        picture = browser.find_element_by_css_selector("#mosaic .zoom img").get_attribute("src")
        get_image = requests.get(picture)
        with open(f"images/{photo_id}.jpg", 'wb') as img_file:
            img_file.write(get_image.content)
        x += 1
        photo_id = (sheet.cell(row=x, column=4)).value

    browser.close()
    browser.quit()
except Exception as ex:
    print(ex)
    browser.close()
    browser.quit()



