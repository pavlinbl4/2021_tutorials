# из отчета TASS извлекает код фото и сумму
import openpyxl

wb = openpyxl.load_workbook('/Users/evgeniy/Downloads/xlsx/Павленко.xlsx')
sheet = wb.active
photos = {}
x = 7
photo_id = (sheet.cell(row=x, column=4)).value
money = (sheet.cell(row=x, column=6)).value
while photo_id != None:
    photos.setdefault(photo_id,[])
    photos[photo_id].append(round(money,3))
    x += 1
    photo_id = (sheet.cell(row=x, column=4)).value
    money = (sheet.cell(row=x, column=6)).value
for i in photos:
    print(f"{i} - {photos[i]}")
