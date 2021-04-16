# из отчета TASS извлекает код фото и сумму
import openpyxl
# file_way = "/Volumes/big4photo/Downloads/xls/Павленко.xlsx"
def get_publications(file_way):
    wb = openpyxl.load_workbook(file_way)
    sheet = wb.active
    photos = {}
    x = 7
    photo_id = (sheet.cell(row=x, column=4)).value
    money = (sheet.cell(row=x, column=6)).value
    while photo_id != None:
        photos.setdefault(photo_id,[])
        photos[photo_id].append(round(money,3))
        x += 1
        photo_id = (sheet.cell(row=x, column=4)).value
        money = (sheet.cell(row=x, column=6)).value
    # for i in photos:
    #     print(f"{i} - {photos[i]}")
    return photos

# get_publications(file_way)
