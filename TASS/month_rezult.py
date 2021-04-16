"""
информация за месяц - сколько какое фото продано
добавляю вкладку в общий файл
"""

import work_with_XLXS
import report_date_from_file
from openpyxl import workbook
from openpyxl import load_workbook


file = "/Volumes/big4photo/Documents/Tass_total_report.xlsx"
file_way = '/Volumes/big4photo/Downloads/Павленко.xlsx'
report_date = report_date_from_file.get_report_date(file_way)
photos = work_with_XLXS.get_publications(file_way)

wb = load_workbook(filename=file, read_only=False)
# ws = wb.active
ws_month_number = wb.create_sheet(report_date, 0)
ws_month_number.cell(row =2,column=1).value = "     photo_id"
ws_month_number.cell(row =2,column=2).value = "     income"
ws_month_number.cell(row =2,column=3).value = "     sold times"

kkeys =[i for i in photos.keys()]
for i in range(len(photos)):
    ws_month_number.cell(row=3 + i, column=1).value = kkeys[i]
    ws_month_number.cell(row=3 + i, column=2).value = sum(photos[kkeys[i]])
    ws_month_number.cell(row=3 + i, column=3).value = len(photos[kkeys[i]])



wb.save(file)
wb.close()