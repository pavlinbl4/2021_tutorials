"""
бот через заданный интервал проверяет
количество моих фото на сайте и заносит данные в таблицу

необходимые улучшение - добавлять данные, только если количество снимков изменилось
"""

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import datetime
import time
import telebot
from auth_data import token

count = 10 # количество запросов на сайт
while count != 0:
    url = 'https://www.tassphoto.com/ru/asset/fullTextSearch/search/%D1%81%D0%B5%D0%BC%D0%B5%D0%BD+%D0%BB%D0%B8%D1%85%D0%BE%D0%B4%D0%B5%D0%B5%D0%B2/page/1'
    headers = {
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4200.0 Iron Safari/537.36",
        "accept": "*/*"}
    req = requests.get(url, headers=headers)
    src = req.text
    soup = BeautifulSoup(src, 'lxml')
    photos_count = str(soup.select(".result-counter#nb-result"))[42:47]

    book = load_workbook(
        "/Volumes/big4photo/Documents/TASS/Tass_data/TASS_photos.xlsx")  # открываю существующий файл для дозаписи
    sheet = book.active

    # sheet.cell(row=1, column=1).value = "Дата запроса"
    # sheet.cell(row=1, column=2).value = "Количество снимков"

    if sheet.cell(row=1, column=3).value != None:
        line_count = sheet.cell(row=1, column=3).value
    else:
        line_count = 3

    if count <= 50:
        sheet.cell(row=line_count, column=1).value = datetime.now().strftime('%Y-%m-%d %H:%M')
        sheet.cell(row=line_count, column=2).value = photos_count
        sheet.cell(row=1, column=3).value = line_count + 1

    book.save("/Volumes/big4photo/Documents/TASS/Tass_data/TASS_photos.xlsx")
    book.close()

    result = f"{datetime.now().strftime('%Y-%m-%d %H:%M')}\nОпубликованно фото: {photos_count}"

    bot = telebot.TeleBot(token)
    # text = "просто сообщение от бота"
    bot.send_message(chat_id=187597961, text=result)


    print(result)

    time.sleep(1800)
    count -= 1
