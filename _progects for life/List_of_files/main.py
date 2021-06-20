"""
задача - есть спискок файлов в директории, сформировать exel файл с их именами,
даннай файл создан для работы с итоговыми изображениями Нового проспекта
"""

import openpyxl
import os

way_to_files = "/Users/evgeniy/PycharmProjects/2021_tutorials/Best_OFF/folders_NP/май 2021/"


def test_way_to_files(way_to_files):
    choise = input(f"Будут обработаны файлы в папке {way_to_files} \n  enter  Y or N ")
    choise.islower()
    if choise == 'n':
        way_to_files = input("Enter way to files")
        test_way_to_files(way_to_files)
    elif choise == 'y':
        pass
    else:
        print("PLEASE ENTER 'Y' OT 'N'")
        test_way_to_files(way_to_files)
    return way_to_files


way_to_files = test_way_to_files(way_to_files)



book = openpyxl.Workbook()
sheet = book.active
book.title = "New Title"
sheet.cell(row=1, column=1).value = "Дата публикации"
sheet.cell(row=1, column=2).value = "Название материала"
sheet.cell(row=1, column=3).value = "Сумма"
names = os.listdir(way_to_files)
row_number = 2
for name in names:
    row_number += 1
    if name[0] != ".":
        sheet.cell(row=row_number, column=1).value = name[:10]
        sheet.cell(row=row_number, column=2).value = name[12:-3]
        sheet.cell(row=row_number, column=3).value = 500

book.save(f"/Volumes/big4photo/Documents/NP_month{name[4:7]}.xlsx")
book.close()
